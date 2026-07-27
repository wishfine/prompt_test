# -*- coding: utf-8 -*-
"""将高中物理两阶段评级 JSONL 生成便于筛选复核的 Excel 工作簿。"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


LEVELS = [f"难度{i}档" for i in range(1, 6)]
LEVEL_INDEX = {level: index + 1 for index, level in enumerate(LEVELS)}

GROUP_COLORS = {
    "标识": "475569",
    "第一阶段": "2563EB",
    "第二阶段": "7C3AED",
    "最终结果": "EA580C",
    "最终分数": "0F766E",
}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            row = json.loads(line)
            if row.get("question_id") is None:
                raise ValueError(f"{path} 第{line_number}行缺少 question_id")
            rows.append(row)
    return rows


def read_labels(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None:
        return {}
    return {
        str(row["question_id"]): row
        for row in read_jsonl(path)
    }


def flatten(value: Any, prefix: str = "") -> dict[str, Any]:
    """递归展开对象；数组保留为JSON文本，避免列数无限扩张。"""
    if not isinstance(value, dict):
        return {prefix: value}
    flattened: dict[str, Any] = {}
    for key, child in value.items():
        child_prefix = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(child, dict):
            flattened.update(flatten(child, child_prefix))
        else:
            flattened[child_prefix] = child
    return flattened


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if isinstance(value, str):
        # Excel 单个单元格最多32767字符。
        return value[:32760]
    return value


def ordered_union(
    flattened_rows: list[dict[str, Any]],
) -> list[str]:
    seen: set[str] = set()
    fields: list[str] = []
    for row in flattened_rows:
        for field in row:
            if field not in seen:
                seen.add(field)
                fields.append(field)
    return fields


def style_header(ws, groups: list[str]) -> None:
    for column, group in enumerate(groups, start=1):
        cell = ws.cell(row=1, column=column)
        cell.fill = PatternFill("solid", fgColor=GROUP_COLORS[group])
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def width_for_header(header: str) -> float:
    if header == "题目ID":
        return 22
    if any(token in header for token in (
        "reason",
        "evidence",
        "analysis",
        "difficulty_source",
        "stage2_error",
    )):
        return 45
    if any(token in header for token in (
        "features",
        "methods",
        "knowledge",
        "corrections",
    )):
        return 26
    return 18


def create_detail_sheet(
    wb: Workbook,
    results: list[dict[str, Any]],
    labels: dict[str, dict[str, Any]],
) -> tuple[list[str], list[str]]:
    ws = wb.active
    ws.title = "评级明细"

    stage1_rows = [
        flatten(row.get("difficulty_rating_stage1") or {})
        for row in results
    ]
    stage2_rows = [
        flatten(row.get("verification") or {})
        for row in results
    ]
    stage1_fields = ordered_union(stage1_rows)
    stage2_fields = ordered_union(stage2_rows)

    final_headers = [
        "GPT5.6复核标签",
        "第一阶段档位",
        "模型建议档位",
        "最终档位",
        "是否一致",
        "档位差",
        "是否一档内",
        "是否严重偏差",
        "最终调整说明",
        "是否需要人工复核",
        "第二阶段状态",
        "Pipeline版本",
    ]
    headers = (
        ["题目ID"]
        + [f"第一阶段.{field}" for field in stage1_fields]
        + [f"第二阶段.{field}" for field in stage2_fields]
        + final_headers
        + ["最终分数"]
    )
    groups = (
        ["标识"]
        + ["第一阶段"] * len(stage1_fields)
        + ["第二阶段"] * len(stage2_fields)
        + ["最终结果"] * len(final_headers)
        + ["最终分数"]
    )
    ws.append(headers)

    for result, stage1_flat, stage2_flat in zip(
        results,
        stage1_rows,
        stage2_rows,
    ):
        question_id = str(result["question_id"])
        label = labels.get(question_id, {}).get(
            "reviewed_difficulty_level",
            "",
        )
        step1_level = result.get("difficulty_level_step1", "")
        final_level = result.get("final_difficulty_level", "")
        gap = (
            LEVEL_INDEX[final_level] - LEVEL_INDEX[label]
            if label in LEVEL_INDEX and final_level in LEVEL_INDEX
            else ""
        )
        verification = result.get("verification") or {}
        final_score = verification.get("reviewed_predicted_accuracy")
        if final_score is None:
            final_score = (
                result.get("difficulty_rating_stage1") or {}
            ).get("predicted_accuracy")
        final_values = [
            label,
            step1_level,
            result.get("model_suggested_level", ""),
            final_level,
            (
                "一致"
                if label and final_level and label == final_level
                else "不一致" if label and final_level else ""
            ),
            gap,
            (
                "是"
                if isinstance(gap, int) and abs(gap) <= 1
                else "否" if isinstance(gap, int) else ""
            ),
            (
                "是"
                if isinstance(gap, int) and abs(gap) >= 2
                else "否" if isinstance(gap, int) else ""
            ),
            result.get("final_adjustment", ""),
            "是" if result.get("needs_manual_review") is True else "否",
            result.get("verification_status", "success"),
            result.get("pipeline_version", ""),
        ]
        ws.append(
            [question_id]
            + [excel_value(stage1_flat.get(field)) for field in stage1_fields]
            + [excel_value(stage2_flat.get(field)) for field in stage2_fields]
            + final_values
            + [final_score if final_score is not None else ""]
        )

    style_header(ws, groups)
    for column, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column)].width = (
            width_for_header(header)
        )
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "@"
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
    score_column = len(headers)
    ws.conditional_formatting.add(
        f"{get_column_letter(score_column)}2:"
        f"{get_column_letter(score_column)}{ws.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="FCA5A5",
            mid_type="num",
            mid_value=58,
            mid_color="FDE68A",
            end_type="num",
            end_value=100,
            end_color="86EFAC",
        ),
    )
    return stage1_fields, stage2_fields


def create_question_sheet(
    wb: Workbook,
    results: list[dict[str, Any]],
    labels: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("题目信息")
    headers = [
        "题目ID",
        "GPT5.6复核标签",
        "题干",
        "选项",
        "解析",
        "题干图片",
        "解析图片",
        "输入充分性",
    ]
    ws.append(headers)
    for result in results:
        question_id = str(result["question_id"])
        quality = result.get("input_quality") or {}
        ws.append([
            question_id,
            labels.get(question_id, {}).get(
                "reviewed_difficulty_level",
                "",
            ),
            excel_value(result.get("stem")),
            excel_value(result.get("options")),
            excel_value(result.get("analysis")),
            excel_value(result.get("stem_image_url")),
            excel_value(result.get("analysis_image_url")),
            quality.get("input_sufficiency", ""),
        ])
    style_header(ws, ["标识"] + ["最终结果"] * (len(headers) - 1))
    widths = [22, 18, 60, 45, 70, 40, 40, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "@"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_summary_sheet(
    wb: Workbook,
    results: list[dict[str, Any]],
    labels: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("汇总")
    truth = Counter()
    prediction = Counter()
    confusion = {
        level: Counter()
        for level in LEVELS
    }
    exact = within_one = severe = evaluated = 0
    for result in results:
        question_id = str(result["question_id"])
        label = labels.get(question_id, {}).get(
            "reviewed_difficulty_level",
        )
        pred = result.get("final_difficulty_level")
        if label not in LEVEL_INDEX or pred not in LEVEL_INDEX:
            continue
        evaluated += 1
        truth[label] += 1
        prediction[pred] += 1
        confusion[label][pred] += 1
        gap = LEVEL_INDEX[pred] - LEVEL_INDEX[label]
        exact += gap == 0
        within_one += abs(gap) <= 1
        severe += abs(gap) >= 2

    ws.append(["指标", "数值"])
    ws.append(["结果题数", len(results)])
    ws.append(["有效评测数", evaluated])
    ws.append(["完全一致率", exact / evaluated if evaluated else None])
    ws.append(["一档内比例", within_one / evaluated if evaluated else None])
    ws.append(["严重偏差数", severe])
    ws["B4"].number_format = "0.00%"
    ws["B5"].number_format = "0.00%"

    start = 9
    ws.cell(start, 1, "档位")
    ws.cell(start, 2, "GPT5.6标签")
    ws.cell(start, 3, "模型预测")
    for offset, level in enumerate(LEVELS, start=1):
        ws.cell(start + offset, 1, level)
        ws.cell(start + offset, 2, truth[level])
        ws.cell(start + offset, 3, prediction[level])

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "标签与模型档位分布"
    chart.y_axis.title = "题目数"
    chart.x_axis.title = "难度档位"
    chart.add_data(
        Reference(
            ws,
            min_col=2,
            max_col=3,
            min_row=start,
            max_row=start + len(LEVELS),
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=1,
            min_row=start + 1,
            max_row=start + len(LEVELS),
        )
    )
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "E2")

    matrix_start = 17
    ws.cell(matrix_start, 1, "真实\\预测")
    for column, level in enumerate(LEVELS, start=2):
        ws.cell(matrix_start, column, level)
    for row_index, truth_level in enumerate(LEVELS, start=1):
        ws.cell(matrix_start + row_index, 1, truth_level)
        for column_index, pred_level in enumerate(LEVELS, start=2):
            ws.cell(
                matrix_start + row_index,
                column_index,
                confusion[truth_level][pred_level],
            )

    for row in (1, start, matrix_start):
        for cell in ws[row]:
            if cell.value is not None:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=GROUP_COLORS["最终结果"],
                )
                cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 22
    for column in range(2, 7):
        ws.column_dimensions[get_column_letter(column)].width = 16
    ws.sheet_view.showGridLines = False


def create_field_guide(
    wb: Workbook,
    stage1_fields: list[str],
    stage2_fields: list[str],
) -> None:
    ws = wb.create_sheet("字段说明")
    ws.append(["字段组", "Excel列名", "说明"])
    for field in stage1_fields:
        ws.append([
            "第一阶段",
            f"第一阶段.{field}",
            "第一阶段模型输出或程序派生字段",
        ])
    for field in stage2_fields:
        ws.append([
            "第二阶段",
            f"第二阶段.{field}",
            "第二阶段结构审计输出或程序复核字段",
        ])
    ws.append([
        "最终分数",
        "最终分数",
        (
            "第二阶段复核后的乘数后正确率；"
            "第二阶段失败时回退第一阶段乘数后正确率"
        ),
    ])
    style_header(ws, ["标识", "标识", "标识"])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 65
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument("--labels")
    parser.add_argument("--output", required=True)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    results = read_jsonl(Path(args.input))
    labels = read_labels(Path(args.labels) if args.labels else None)
    wb = Workbook()
    stage1_fields, stage2_fields = create_detail_sheet(
        wb,
        results,
        labels,
    )
    create_question_sheet(wb, results, labels)
    create_summary_sheet(wb, results, labels)
    create_field_guide(wb, stage1_fields, stage2_fields)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(json.dumps({
        "input": args.input,
        "rows": len(results),
        "stage1_columns": len(stage1_fields),
        "stage2_columns": len(stage2_fields),
        "output": str(output),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
